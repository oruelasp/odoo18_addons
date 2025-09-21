
# -*- coding: utf-8 -*-
import base64
import io
from odoo import models, fields, api, _
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


class MrpEco(models.Model):
    _inherit = 'mrp.eco'

    # --- Sample Order Attributes ---
    sample_size = fields.Char(string="Talla")
    sample_garment_count = fields.Integer(string="Nº Prendas")
    sample_designer_id = fields.Many2one('hr.employee', string="Diseñador")
    sample_season = fields.Char(string="Temporada")
    sample_sale_month = fields.Selection([
        ('1', 'Enero'), ('2', 'Febrero'), ('3', 'Marzo'), ('4', 'Abril'),
        ('5', 'Mayo'), ('6', 'Junio'), ('7', 'Julio'), ('8', 'Agosto'),
        ('9', 'Septiembre'), ('10', 'Octubre'), ('11', 'Noviembre'), ('12', 'Diciembre'),
    ], string="Mes Venta")
    sample_year = fields.Char(string="Año")
    sample_type = fields.Char(string="Tipo Muestra")
    sample_maker = fields.Many2one(
        'res.partner',
        string="Maquilador",
        domain="[('supplier_rank', '>', 0)]"
    )
    sample_circuit_number = fields.Char(string="Nº Circuito", required=True)
    sample_delivery_date = fields.Date(string="Fecha de Entrega Requerida")

    # --- Report Method ---
    def action_generate_excel_report(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(_("La librería 'openpyxl' no está instalada. Por favor, ejecute 'pip install openpyxl' en su terminal."))

        bom = self.bom_id
        if not bom:
            raise UserError(_("La Orden de Muestra no tiene una Lista de Materiales (LdM) de referencia en la pestaña 'Información Adicional'. Por favor, asígnela."))

        # Create workbook and select active sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # --- STYLES ---
        title_font = Font(bold=True, size=14)
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Number format style
        number_style = NamedStyle(name='number_style', number_format='0.00')
        workbook.add_named_style(number_style)

        # --- HELPERS ---
        def set_cell(cell, value, font=None, alignment=None, border=None, style=None):
            sheet[cell] = value
            if font: sheet[cell].font = font
            if alignment: sheet[cell].alignment = alignment
            if border: sheet[cell].border = border
            if style: sheet[cell].style = style
        
        # --- HEADER ---
        sheet.merge_cells('H3:K3')
        set_cell('H3', 'Receta Desarrollo de Prenda', font=title_font, alignment=center_align)

        # --- GENERAL DATA (LEFT) ---
        set_cell('B8', 'Orden de Muestra:', font=bold_font)
        set_cell('C8', f"{self.name} {self.sample_circuit_number or ''}")
        set_cell('B9', 'Marcas:', font=bold_font)
        set_cell('C9', self.product_tmpl_id.attribute_line_ids.filtered(lambda l: l.attribute_id.name.upper() == 'MARCA').value_ids[:1].name or '')
        set_cell('B10', 'Tipo de Prenda:', font=bold_font)
        set_cell('C10', self.product_tmpl_id.attribute_line_ids.filtered(lambda l: l.attribute_id.name.upper() == 'TIPO DE PRENDA').value_ids[:1].name or '')
        set_cell('B11', 'Clasificación:', font=bold_font)
        set_cell('C11', self.product_tmpl_id.attribute_line_ids.filtered(lambda l: l.attribute_id.name.upper() == 'GRUPO').value_ids[:1].name or '')
        set_cell('B12', 'Año Venta:', font=bold_font)
        set_cell('C12', self.sample_year or '')
        set_cell('B13', 'Tipo de Muestra:', font=bold_font)
        set_cell('C13', self.sample_type or '')

        # --- GENERAL DATA (CENTER) ---
        set_cell('G8', 'Nº Circuito:', font=bold_font)
        set_cell('H8', self.sample_circuit_number or '')
        set_cell('G9', 'Talla:', font=bold_font)
        set_cell('H9', self.sample_size or '')
        set_cell('G10', 'Nº Prendas:', font=bold_font)
        set_cell('H10', self.sample_garment_count or '')
        set_cell('G11', 'Temporada:', font=bold_font)
        set_cell('H11', self.sample_season or '')
        set_cell('G12', 'Diseñador:', font=bold_font)
        set_cell('H12', self.sample_designer_id.name or '')
        set_cell('G13', 'Maquilador:', font=bold_font)
        set_cell('H13', self.sample_maker.name or '')
        set_cell('G14', 'Impreso por:', font=bold_font)
        set_cell('H14', self.env.user.name)

        # --- DYNAMIC ATTRIBUTES & DATES (RIGHT) ---
        row = 8
        for attr_line in self.product_tmpl_id.attribute_line_ids:
            attr_name = attr_line.attribute_id.name
            if attr_name.upper() not in ['MARCA', 'TIPO DE PRENDA', 'GRUPO']:
                set_cell(f'J{row}', f'{attr_name}:', font=bold_font)
                set_cell(f'K{row}', attr_line.value_ids[:1].name or '')
                row += 1
        
        set_cell(f'J{row}', 'Fecha:', font=bold_font)
        set_cell(f'K{row}', fields.Date.today(), alignment=right_align)

        # --- RUTA DE PROCESO ---
        row = 16
        set_cell(f'B{row}', 'RUTA DE PROCESO', font=bold_font)
        row += 1
        set_cell(f'B{row}', 'Proceso', font=bold_font, border=thin_border)
        set_cell(f'C{row}', 'Proveedor/Taller', font=bold_font, border=thin_border)
        set_cell(f'D{row}', 'Importe', font=bold_font, border=thin_border)
        set_cell(f'E{row}', 'Observaciones', font=bold_font, border=thin_border)
        
        mol_total = 0
        if bom.operation_ids:
            fresh_operations = self.env['mrp.routing.workcenter'].search([('bom_id', '=', self.bom_id.id)])
            for op in fresh_operations:
                row += 1
                set_cell(f'B{row}', op.name, border=thin_border)
                set_cell(f'C{row}', op.workcenter_id.name, border=thin_border)
                set_cell(f'D{row}', op.cost, border=thin_border, style=number_style)
                set_cell(f'E{row}', op.note, border=thin_border)
                mol_total += op.cost
        
        row += 1
        set_cell(f'C{row}', 'Sub-Total', font=bold_font, alignment=right_align)
        set_cell(f'D{row}', mol_total, font=bold_font, style=number_style)
        
        # --- COMPONENT SECTIONS (DYNAMIC) ---
        row += 2
        
        def get_root_category_name(category):
            # This robust method uses the pre-computed parent_path to find the root.
            if category.parent_path:
                try:
                    # parent_path is like '1/15/42/'. We want the first ID.
                    root_id = int(category.parent_path.split('/')[0])
                    # Search for the category to get a fresh record.
                    root_category = self.env['product.category'].search([('id', '=', root_id)], limit=1)
                    if root_category:
                        return root_category.name
                except (ValueError, IndexError):
                    # Fallback in case parent_path is malformed
                    pass
            # Fallback to the category's own name if no parent path or if something goes wrong.
            return category.name

        components_by_category = {}
        for line in bom.bom_line_ids:
            if line.product_id and line.product_id.categ_id:
                category_name = get_root_category_name(line.product_id.categ_id) or 'Sin Categoría'
            else:
                category_name = 'Sin Categoría'
            
            if category_name not in components_by_category:
                components_by_category[category_name] = []
            components_by_category[category_name].append(line)

        components_total = 0
        for category_name in sorted(components_by_category.keys()):
            set_cell(f'B{row}', category_name.upper(), font=bold_font)
            row += 1
            headers = ['Código', 'Descripción', 'U/M', 'Consumo', 'Precio S/.', 'Importe', 'Proveedor', 'Observaciones']
            col_letters = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
            for col, header in zip(col_letters, headers):
                set_cell(f'{col}{row}', header, font=bold_font, border=thin_border)

            section_subtotal = 0
            for line in components_by_category[category_name]:
                row += 1
                price = line.product_id.standard_price
                consumption = line.product_qty
                amount = price * consumption
                supplier = line.product_id.seller_ids[:1].partner_id.name or ''

                set_cell(f'B{row}', line.product_id.default_code, border=thin_border)
                set_cell(f'C{row}', line.product_id.name, border=thin_border)
                set_cell(f'D{row}', line.product_uom_id.name, border=thin_border)
                set_cell(f'E{row}', consumption, border=thin_border, style=number_style)
                set_cell(f'F{row}', price, border=thin_border, style=number_style)
                set_cell(f'G{row}', amount, border=thin_border, style=number_style)
                set_cell(f'H{row}', supplier, border=thin_border)
                set_cell(f'I{row}', '', border=thin_border)
                section_subtotal += amount
            
            row += 1
            set_cell(f'F{row}', 'Sub-Total', font=bold_font, alignment=right_align)
            set_cell(f'G{row}', section_subtotal, font=bold_font, style=number_style)
            components_total += section_subtotal
            row += 1 # Reduced space
        
        # --- FINAL TOTALS ---
        row += 1
        set_cell(f'F{row}', 'Total:', font=bold_font, alignment=right_align)
        set_cell(f'G{row}', components_total, font=bold_font, style=number_style)
        row += 1
        set_cell(f'F{row}', 'M.O.L.:', font=bold_font, alignment=right_align)
        set_cell(f'G{row}', mol_total, font=bold_font, style=number_style)
        row += 1
        set_cell(f'F{row}', 'COSTO:', font=bold_font, alignment=right_align)
        set_cell(f'G{row}', components_total + mol_total, font=bold_font, style=number_style)
        row += 1
        set_cell(f'F{row}', 'PVP:', font=bold_font, alignment=right_align)
        set_cell(f'G{row}', self.product_tmpl_id.list_price, font=bold_font, style=number_style)
        row += 1
        set_cell(f'F{row}', 'PVP Ripley:', font=bold_font, alignment=right_align)
        set_cell(f'G{row}', '', font=bold_font)

        # --- ADJUST COLUMN WIDTHS ---
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        # --- SAVE AND DOWNLOAD ---
        fp = io.BytesIO()
        workbook.save(fp)
        fp.seek(0)
        file_data = base64.b64encode(fp.read())
        fp.close()

        attachment_name = f'Receta_Desarrollo_{self.name.replace("/", "_")}.xlsx'
        attachment = self.env['ir.attachment'].create({
            'name': attachment_name,
            'datas': file_data,
            'res_model': 'mrp.eco',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
